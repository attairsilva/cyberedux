#------------------------------------------------------------------------------

import os
from dotenv import load_dotenv
from groq import Groq
import json
import telebot
from openpyxl import load_workbook


load_dotenv() 
# print(os.getenv("TOKEN_BOT_TELEGRAM")) 

bot = telebot.TeleBot(os.getenv('TOKEN_BOT_TELEGRAM'))
cliente_groq = Groq(api_key=os.getenv("TOKEN_API_GROQ"))


cardapio = {}


def carregar_cardapio():
    try: 
        excel = load_workbook('cardapio.xlsx')
        planilha=excel['Sheet']
        total_linhas=planilha.max_row
        for linhas in range(1,total_linhas+1):
            id = planilha.cell(row=linhas, column=1).value
            nome = planilha.cell(row=linhas, column=2).value
            preco = planilha.cell(row=linhas, column=3).value
            cardapio[id] = {'nome': nome, 'preco': preco}
        print('Cardapio carregados em lista')
        
    except Exception as e:
        print(f'Erro ao carregar excel de cardapio: {e} ')

def carregar_prompt():
    try: 
        with open('prompt.txt', 'r') as arquivo_prompt:
            prompt = arquivo_prompt.read()
        return prompt
    except Exception as e:
        print (f'Erro ao carregar prompt {e}')

carregar_cardapio()

# carregando cardapio no historico junto ao prompt
historico_mensagens = [
    {
        "role": "system",
        "content": f"O cardápio atual é:\n{cardapio}\n\n{carregar_prompt()}"
    }
]

def processar_mensagem(mensagem):
    
    historico_mensagens.append({
        "role": "user",
        "content": mensagem 
    })
      
    try:
        completion = cliente_groq.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=historico_mensagens,
            temperature=1,
            max_completion_tokens=1024,
            top_p=1,
            stream=False,
            stop=None,
            response_format={'type': 'json_object'}
        )
    except:
        historico_mensagens.pop()
        return {'intencao': 0, 'resposta': 'Erro servidor Groq'}

    resposta_ia = completion.choices[0].message
    resposta_dados = json.loads(resposta_ia.content)

    historico_mensagens.append(resposta_ia)

    return resposta_dados


#------------------------------------------------------------------------------
# Telegram
#------------------------------------------------------------------------------

# Responde ao comando /start
@bot.message_handler(commands=['start'])#, 'help'])
def mensagem_inicial(message):
    nome = bot.get_my_name().name
    bot.reply_to(message, f'Ola! Sou seu agente de atendimento {nome}!')

# Responde ao comando /sair
@bot.message_handler(commands=['sair'])
def sair(message):
    bot.reply_to(message, 'Ok, :(')
    bot.stop_polling()

# Responde as demais mensagens
@bot.message_handler(func=lambda message: True)
def assistente(message):
    
    dados = processar_mensagem(message.text)

    resposta = ""
    
    if dados['intencao'] == 1:
        
        resposta  = "Segue o nosso cardápio:\n" 
        for item, dados_item in cardapio.items():
            preco = dados_item.get('preco')
            if preco is not None:
                resposta += f"{dados_item['nome']} - R$ {preco:.2f}\n"
               

    elif dados['intencao'] == 2:
        
            resposta = f'Anotando o pedido, {dados['resposta']}'
       
    else:
        
        #resposta = f'Comando fora de contexto, {dados['resposta']}'
        resposta = f'{dados['resposta']}'

    bot.send_message(message.chat.id, resposta)
    print(f'Recebido do usuario: {message.text}')


# roda o TeleBot
bot.infinity_polling()
