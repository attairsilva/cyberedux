import requests
from openpyxl import load_workbook
from docxtpl import DocxTemplate

# Função para obter a lista de preços
# ------------------------------------------------------------------------------
def obter_precos(cod_ativo):
    url = "https://statusinvest.com.br/acao/tickerprice"

    params = {
        "ticker" : cod_ativo,
        "type": 1
    }

    headers = {
        "User-Agent": "Mozilla/5.0"
    }

    http_req = requests.get(url, params=params, headers=headers)

    dados = http_req.json()
    return dados[0]["prices"]

# ------------------------------------------------------------------------------


def salvarExcel(lista_precos):

    excel = load_workbook('precos.xlsx')
    planilha = excel['dados']
   
    linha=2
    for elemento in lista_precos:
        planilha.cell(linha,1).value = elemento['date']
        planilha.cell(linha,2).value = elemento['price']
        linha+=1
    
    excel.save('preco-alterado.xlsx')

def salvarDocx(lista_precos, ticker):
    template = DocxTemplate('template_precos.docx')
    contexto = {
        'ativo' :  ticker,
        'precos': lista_precos
    }

    template.render(contexto)
    template.save('template_precos_alterado.docx')

ticker = input("Digite o ticker do ativo: ")
lista_precos=obter_precos(ticker)
salvarExcel(lista_precos)
salvarDocx(lista_precos,ticker)
