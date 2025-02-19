import requests
from openpyxl import load_workbook
from docxtpl import DocxTemplate

# Função para obter a lista de preços
# ------------------------------------------------------------------------------
def obter_precos(cod_ativo):
    url = "https://statusinvest.com.br/acao/tickerprice"

    params = {
        "ticker" : cod_ativo,
        "type": 1
    }

    headers = {
        "User-Agent": "Mozilla/5.0"
    }

    http_req = requests.get(url, params=params, headers=headers)

    dados = http_req.json()
    return dados[0]["prices"]

# ------------------------------------------------------------------------------

def salvar_excel(precos):
    excel = load_workbook('precos.xlsx')
    planilha = excel['dados']

    # modo 1
    linha = 2
    for dado in precos:
        planilha.cell(linha,1).value = dado['date']
        planilha.cell(linha,2).value = dado['price']
        linha += 1

    # # modo 2
    # for linha, dado in enumerate(precos, 2):
    #     planilha.cell(linha,1).value = dado['date']
    #     planilha.cell(linha,2).value = dado['price']
    
    # # modo 3
    # for idx in range(0, len(precos)):
    #     planilha.cell(idx+2,1).value = precos[idx]['date']
    #     planilha.cell(idx+2,2).value = precos[idx]['price']

    excel.save('precos_alterado.xlsx')

def salvar_docx(ticker, precos):
    doc = DocxTemplate('template_precos.docx')

    contexto = {
        'precos': precos,
        'ativo': ticker
    }

    doc.render(contexto)

    doc.save('relatorio.docx')

ticker = input('Código do ativo: ')
lista_precos = obter_precos(ticker)

salvar_excel(lista_precos)
salvar_docx(ticker, lista_precos)
