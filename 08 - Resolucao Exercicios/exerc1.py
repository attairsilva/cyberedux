"""
Faça um programa para disparar emails de confirmação de inscrição.
Crie ao menos 3 emails temporários usando https://mail.tm,
Coloque-os em uma planilha e mande os emails de forma automática.

--- Os endereços de email devem ser lidos da planilha.
"""

from smtplib import SMTP_SSL
from email.mime.text import MIMEText
from openpyxl import load_workbook
 
import json

def carregar_config():
# carrega as configurações do arquivo config.json
    with open('config.json', 'r') as arq:
        config = json.load(arq)

    return config

def login_email(config):
    con = SMTP_SSL(config['servidor'][0], config['servidor'][1])
    con.login(config['usuario'], config['senha'])

    return con

def enviar_email(con, destinatario, remetente, assunto, corpo):
    msg = MIMEText(corpo)
    msg['Subject'] = assunto
    msg['From'] = remetente

    con.sendmail(remetente, destinatario, msg.as_string())

def enviar_emails_confirmacao():
    config = carregar_config()
    con = login_email(config)

    excel = load_workbook('inscricoes.xlsx')
    planilha = excel.active
    total_linhas = planilha.max_row

    for linha in range(2, total_linhas+1):
        insc = planilha.cell(linha, 1).value
        nome = planilha.cell(linha, 2).value 
        email = planilha.cell(linha, 3).value
    
        assunto = 'Inscrição confirmada!'
        corpo_email = f'Parabéns {nome}, sua inscrição de Nº {insc} está confirmada!'
        enviar_email(con, email, config['usuario'], assunto, corpo_email)

enviar_emails_confirmacao()

