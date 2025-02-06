import json  
from openpyxl import Workbook, load_workbook



# Importa o módulo JSON para manipular arquivos no formato JSON
excel = Workbook()

def cadastrar_pessoa(dicionario):
    #funcao cadastra pessoa no dicionaio
    cpf=input('Informe o seu CPF: ')
    nome=input('Informe seu nome: ') 
    pessoa = []
    pessoa=nome
    dicionario[cpf]={ "nome": pessoa}

def buscar_pessoas(pessoas):
    #busca pessoa no dicionario
    consulta = int(input('Informe o CPF da pessoa: '))
    cont=0
    for chave, valor in pessoas.items():
        if int(chave) == consulta:
            cont+=1
            print(f'{chave} - Nome: {valor['nome']} ')
        if cont == 0:
            print('Não encontrado!')

def listar_pessoas(pessoas):
    # lista pessoas do dicionario
    for cpf,pessoa in pessoas.items():
                print(f'{cpf} - Nome: {pessoa['nome']} ')

def salvar_pessoas_json(pessoas):
    # Função para salvar os cadastros em um arquivo JSON
    with open("cadastros.json", "w") as arquivo:  
            # Abre o arquivo para escrita
            json.dump(pessoas, arquivo, indent=4)  
             # Salva a lista de cadastros no arquivo
            print("Cadastros salvos com sucesso no json!")  
            # Confirmação de salvamento

def salvar_pessoas_excel(excel,pessoas):
    # Função para salvar na planilha
    # planilha = excel.create_sheet('cadastro')
    
    planilha = excel['Sheet']
    
    planilha['A1'] = 'Nome'
    planilha['B1'] = 'CPF'
    
    linha = 2
    
    for chave, valor in pessoas.items():
        nome = valor['nome']
        planilha.cell(row=linha, column=1).value=nome
        planilha.cell(row=linha, column=2).value=chave
        linha += 1
    try:   
     excel.save('cadastros.xlsx')
    except:
     print("Arquivo a ser salvo esta aberto em outra aplicação!")
     
    print("Cadastros salvos com sucesso no excel!")  
    # Confirmação de salvamento

def carregar_pessoas():
    try:
        with open("cadastros.json", "r") as arquivo:
                pessoas = json.load(arquivo)
                print("Carregados o arquivo JSON com sucesso!")
                return pessoas  # Retorna os dados carregados
    except:
        pessoas=[]


def carregar_excel():
    try: 
        excel = load_workbook('cadastros.xlsx')
        planilha=excel['Sheet']
        total_linhas=planilha.max_row
        for linhas in range(2,total_linhas+1):
            nome = planilha.cell(row=linhas, column=1).value
            cpf = planilha.cell(row=linhas, column=2).value
            pessoas[cpf]={ 'nome': nome}
    except:
        print('Dados do excel carregados em lista')

def editar_pessoas(dicionario):
    # Edita o nome de uma pessoa já cadastrada
    cpf = input("Informe o CPF da pessoa que deseja editar: ")
    print(f'= Nome Anterior => {dicionario.get(cpf)}')
    if cpf in dicionario:
        novo_nome = input("Informe o novo nome: ")
        dicionario[cpf]['nome'] = novo_nome
        print("Cadastro atualizado com sucesso!")
    else:
        print("CPF não encontrado!")

def deletar_pessoas(dicionario):
    # Remove uma pessoa do dicionário
    cpf = input("Informe o CPF da pessoa que deseja excluir: ")
    if cpf in dicionario:
        dicionario.pop(cpf)
        print("Cadastro removido com sucesso!")
    else:
        print("CPF não encontrado!")

pessoas=carregar_pessoas()

while True:
    
    print("\nMenu de Opções:")
    print("1. Incluir pessoa")  
    print("2. Listar cadastros de pessoas")  
    print("3. Editar Pessoa")  
    print("4. Deletar Pessoa") 
    print("5. Buscar pessoa no dicionario") 
    print("6. Carregar do Excel") 
    print("7. Salvar no Excel") 
    print('8. Sair')
    print('==================================')
    operacao = int(input('Informe o numero da opção: '))
    if operacao == 1:
        cadastrar_pessoa(pessoas)
    elif operacao == 2:
        listar_pessoas(pessoas)
    elif operacao == 3:
        editar_pessoas(pessoas)
    elif operacao == 4:
        deletar_pessoas(pessoas)
    elif operacao == 5:
        buscar_pessoas(pessoas) 
    elif operacao == 6:
        carregar_excel()
    elif operacao == 7:
        salvar_pessoas_excel(excel,pessoas)
    elif operacao == 8:
        salvar_pessoas_json(pessoas)
        break 
    
    else:
        print('Opção inválida!')
