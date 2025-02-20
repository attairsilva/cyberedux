from openpyxl import Workbook

excel = Workbook()

planilha = excel.create_sheet('cadastro')

print(excel.sheetnames)
planilha['A1']=10
planilha.cell(row=1, column=2).value=11

print(planilha['A1'].value)
print(planilha['B1'].value)

excel.save('exemplo1.xlsx')
