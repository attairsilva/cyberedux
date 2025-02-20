from openpyxl import Workbook
# Inicia o arquivo Excel
arquivo_excel = Workbook()
# Cria a planilha 'dados'
arquivo_excel.create_sheet('dados')
# Seleciona a planilha dados
planilha = arquivo_excel['dados']
# altera a célula 1,1 (A1) para o valor 'Teste!'
planilha.cell(1,1).value = 'Teste!'
# Salva o arquivo
arquivo_excel.save('cadastro2.xlsx')
