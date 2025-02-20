'''
Faça um programa de cadastro de pessoas. Esse programa deve ter um menu com as seguintes opções:
1. Cadastrar pessoa
2. Listar cadastros
3. Editar cadastro
4. Deletar cadastro
5. Procurar cadastro
0. Sair

*O programa deve abrir o arquivo JSON e carregar seus cadastros ao iniciar.

- Na opção 1, o usuário insere os dados da pessoa (nome e cpf). A pessoa é, então, registrada em uma tabela
formada por coleções de dados.

-Na opção 2, o programa mostra todos os cadastros no terminal.

- Na opção 3, o usuário insere qual CPF quer alterar e quais os novos dados para aquele cadastro (nome) e
então o programa altera os dados. Caso o CPF não existir, o programa deve notificar que o CPF não existe.

- Na opção 4, o usuário insere qual o CPF deseja deletar do cadastro e então o programa remove o cadastro
com aquele CPF. Caso o CPF não existir, o programa deve notificar que o CPF não existe.

- Na opção 5, o usuário insere qual CPF deseja buscar e então o programa busca o cadastro pelo CPF e mostra
os dados daquele cadastro. Caso o CPF não existir, o programa deve notificar que o CPF não existe.

- Na opção 0, o programa deve salvar o cadastro em um arquivo JSON e então finalizar

### MODO 1
cadastro = {
    '000.000.000-00' : 'Fulano',
    '000.000.000-01' : 'Ciclano', 
}

cadastro[cpf] = nome

### MODO 2

cadastro = [
    ['000.000.000-00', 'Fulano'],
    ['000.000.000-01', 'Ciclano'],
]

cadastro.append([cpf, nome])

### MODO 3

cadastro = {
    '000.000.000-00' : { 'nome' : 'Fulano' },
    '000.000.000-01' : { 'nome' : 'Ciclano' }, 
}

cadastro[cpf] = { 'nome' : nome }


cpf = '000.000.000-00'
pessoa = { 'nome' : 'Fulano' },

cpf = '000.000.000-01'
pessoa = { 'nome' : 'Ciclano' }, 


'''

from openpyxl import load_workbook, Workbook

cadastro = {}

def cadastrar_pessoa():
    nome = input('Nome: ')
    cpf = input('CPF: ')
    
    # modo 1
    # pessoa = {}
    # pessoa['nome'] = nome
    # cadastro[cpf] = pessoa

    # modo 2
    if cpf in cadastro:
        print('Cadastro já existe!!!')
    else:
        cadastro[cpf] = { 'nome' : nome }

def listar_cadastro():
    # modo 1
    # for chave, valor in dicionario.items()
    for cpf, pessoa in cadastro.items():
        print(f'{cpf} - {pessoa['nome']}')

    # modo 2
    # for chave in cadastro:
    #     print(f'Cadastro: {cadastro[chave]} - {cadastro[chave]['nome']}')

def editar_cadastro():
    cpf = input('Qual CPF deseja alterar? ')

    if cpf in cadastro:
        # edito
        nome = input('Novo nome: ')
        cadastro[cpf]['nome'] = nome
    else:
        print('CPF não cadastrado!')

def deletar_cadastro():
    cpf = input('Qual CPF deseja deletar? ')

    if cpf in cadastro:
        # deleto
        cadastro.pop(cpf)
    else:
        print('CPF não cadastrado!')

def procurar_cadastro():
    cpf = input('Qual CPF deseja procurar? ')

    if cpf in cadastro:
        # busca
        print(f'{cpf} - {cadastro[cpf]['nome']}')
    else:
        print('CPF não cadastrado!')


def carregar_excel():
    arquivo_excel = load_workbook('cadastro.xlsx')
    planilha = arquivo_excel['dados']

    total_linhas = planilha.max_row

    # Lê as linhas da planilha
    for linha in range(2, total_linhas+1):
        nome = planilha.cell(linha, 1).value
        cpf = planilha.cell(linha, 2).value

        cpf = cpf.strip()

        cadastro[cpf] = { 'nome' : nome }

def salvar():
    # arquivo_excel = Workbook()
    arquivo_excel = load_workbook('cadastro.xlsx')

    planilha = arquivo_excel['dados']

    linha = 2
    for cpf, pessoa in cadastro.items():
        nome = pessoa['nome']

        planilha.cell(linha, 1).value = nome
        planilha.cell(linha, 2).value = cpf
        
        linha += 1
    
    arquivo_excel.save('cadastro.xlsx')



def menu():
    carregar_excel()

    while True:
        print('1. Cadastrar pessoa')
        print('2. Listar cadastros')
        print('3. Editar cadastro')
        print('4. Deletar cadastro')
        print('5. Procurar cadastro')
        print('0. Sair')

        opc = input('Digite a opção: ')

        if opc == '1':
            # cadastrar pessoa
            cadastrar_pessoa()
        elif opc == '2':
            # listar cadastro
            listar_cadastro()
        elif opc == '3':
            # editar cadastro
            editar_cadastro()
        elif opc == '4':
            # deletar cadastro
            deletar_cadastro()
        elif opc == '5':
            # procurar cadastro
            procurar_cadastro()
        elif opc == '0':
            # salvar e sair
            salvar()
            break
        else:
            print(f'Opção inválida: {opc}')

menu()
