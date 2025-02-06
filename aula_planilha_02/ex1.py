from openpyxl import load_workbook
# abrir arquivo Excel
arquivo_excel = load_workbook('cadastro.xlsx')
# Nome das planilhas
print(arquivo_excel.sheetnames)
# Planilha 'dados'
planilha = arquivo_excel['dados']

# O valor das colunas A1 e B1
print(planilha['A1'].value)
print(planilha['B1'].value)

# O total de linhas preenchidas
total_linhas = planilha.max_row

# Lê as linhas da planilha
for linha in range(2, total_linhas+1):
    nome = planilha.cell(linha, 1).value
    cpf = planilha.cell(linha, 2).value

    print(f'{nome} - {cpf}')

planilha.cell(11, 1).value = planilha.cell(11, 1).value.upper()

arquivo_excel.save('cadastro.xlsx')



