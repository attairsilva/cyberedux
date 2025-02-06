from openpyxl import Workbook, load_workbook

carros = [
         ['AVX-0101','Fiat','Toro','Branco'],
         ['TRF-2328','Nissan','Leaf','Prata'],
         ['AVX-3891','Chevrolet','Primas','Branco']
]

try: 
             
    arquivo_excel = load_workbook('carros.xlsx')
    planilha = arquivo_excel['carros']
    
except:
    arquivo_excel = Workbook()
    arquivo_excel.create_sheet('carros')
    # Seleciona a planilha dados
    planilha = arquivo_excel['carros']
    # altera a célula 1,1 (A1) para o valor 'Teste!'
    planilha.cell(1,1).value = 'PLACA'
    planilha.cell(1,2).value = 'MARCA'
    planilha.cell(1,3).value = 'MODELO'
    planilha.cell(1,4).value = 'COR'
    # Salva o arquivo
    arquivo_excel.save('carros.xlsx')

total_linhas = planilha.max_row

linha = total_linhas + 1

for valor in carros:
    coluna=0
    for chave in valor:
         coluna+=1
         planilha.cell(row=linha, column=coluna).value=chave
    linha+=1
         
try:   
    arquivo_excel.save('carros.xlsx')
    print("Arquivo salvo!")
except:
    print("Arquivo a ser salvo esta aberto em outra aplicação!")
    
