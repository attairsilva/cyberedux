from smtplib import SMTP_SSL
from email.mime.text import MIMEText
from openpyxl import load_workbook
import json
# Site que cria e-mails temporarios: https://mail.tm/

 # carrega as configurações do arquivo config.json
with open('config.json', 'r') as arq:
    config = json.load(arq)
servidor = config['servidor']
usuario = config['usuario']
senha = config['senha']

#envia mensagem
def enviaMensagem(assunto,destinatario,corpo_email,con):
    msg = MIMEText(corpo_email, 'html')
    msg['From'] = usuario 
    msg['Subject'] = assunto
    # con é a conexao recebida como parametro
    con.sendmail(usuario, destinatario, msg.as_string())
    
    if con:
        print(f'Email enviado para {destinatario}!')
    else:
        print(f'[Erro] ao enviar para {destinatario}')
    
# funcao carrega dados e envia email
def enviar_excel_email():
   
        con = SMTP_SSL(servidor[0], servidor[1])
        con.login(usuario, senha)
   
        excel = load_workbook('inscritos.xlsx')
        planilha=excel['Sheet']
        
        total_linhas=planilha.max_row
        for linhas in range(2,total_linhas+1):
                nome = planilha.cell(row=linhas, column=1).value
                email = planilha.cell(row=linhas, column=2).value
                corpo_email=f"Prezado Senhor,<br><b>{nome}</b> <p>Queremos lhe dar as boas vinda.</p> <p> Rececebemos o seu pedido de inscrição.</p> <p>Atenciosamente</p><b>Attair B Silva</b>"
                assunto=f"[Aviso de Inscrição] Seja bem vindo {nome}!"
 
                enviaMensagem(assunto,email,corpo_email,con)
                
enviar_excel_email()